"""
Data Structure Analyzer for Oil & Gas Files
Analyzes actual file structures to understand real data formats
"""

import os
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import json
import re
from collections import defaultdict, Counter
import mimetypes
import magic
from datetime import datetime

# For specialized file analysis
try:
    import lasio
    import dlisio
    from PIL import Image
    import PyPDF2
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("Some analysis libraries not available. Installing basic ones...")

class DataStructureAnalyzer:
    """Analyzes actual file structures to understand data formats"""
    
    def __init__(self, data_path: str):
        self.data_path = Path(data_path)
        self.analysis_results = {}
        self.file_inventory = {}
        
    def analyze_all_files(self) -> Dict[str, Any]:
        """Analyze all files in the data directory"""
        print("🔍 Starting comprehensive file structure analysis...")
        print("=" * 60)
        
        # Step 1: File discovery and inventory
        self._discover_files()
        
        # Step 2: Analyze each file type
        self._analyze_by_file_type()
        
        # Step 3: Generate comprehensive report
        report = self._generate_analysis_report()
        
        # Step 4: Save results
        self._save_analysis_results(report)
        
        return report
    
    def _discover_files(self):
        """Discover all files and create inventory"""
        print("📁 Discovering files...")
        
        file_inventory = defaultdict(list)
        total_files = 0
        total_size = 0
        
        for file_path in self.data_path.rglob('*'):
            if file_path.is_file():
                total_files += 1
                file_size = file_path.stat().st_size
                total_size += file_size
                
                # Get file extension
                ext = file_path.suffix.lower()
                
                # Get MIME type
                try:
                    mime_type = magic.from_file(str(file_path), mime=True)
                except:
                    mime_type = mimetypes.guess_type(str(file_path))[0] or 'unknown'
                
                file_info = {
                    'path': str(file_path),
                    'name': file_path.name,
                    'size_bytes': file_size,
                    'size_mb': file_size / (1024 * 1024),
                    'extension': ext,
                    'mime_type': mime_type,
                    'relative_path': str(file_path.relative_to(self.data_path))
                }
                
                file_inventory[ext].append(file_info)
        
        self.file_inventory = dict(file_inventory)
        
        print(f"✅ Discovered {total_files} files ({total_size / (1024*1024):.1f} MB)")
        print(f"📊 File types found: {list(self.file_inventory.keys())}")
        print()
    
    def _analyze_by_file_type(self):
        """Analyze files grouped by type"""
        print("🔬 Analyzing files by type...")
        
        for ext, files in self.file_inventory.items():
            if not files:
                continue
                
            print(f"\n📋 Analyzing {ext} files ({len(files)} files)...")
            
            if ext == '':
                self._analyze_no_extension_files(files)
            elif ext == '.asc':
                self._analyze_ascii_files(files)
            elif ext == '.las':
                self._analyze_las_files(files)
            elif ext == '.dlis':
                self._analyze_dlis_files(files)
            elif ext == '.segy':
                self._analyze_segy_files(files)
            elif ext == '.pdf':
                self._analyze_pdf_files(files)
            elif ext in ['.xlsx', '.xls']:
                self._analyze_excel_files(files)
            elif ext in ['.png', '.jpg', '.jpeg', '.tiff', '.bmp']:
                self._analyze_image_files(files)
            elif ext == '.txt':
                self._analyze_text_files(files)
            elif ext == '.dat':
                self._analyze_dat_files(files)
            else:
                self._analyze_unknown_files(files, ext)
    
    def _analyze_ascii_files(self, files: List[Dict]):
        """Analyze ASCII files structure"""
        print("  📝 Analyzing ASCII file structures...")
        
        ascii_analysis = {
            'file_type': 'ascii',
            'total_files': len(files),
            'file_details': [],
            'common_patterns': defaultdict(int),
            'header_patterns': [],
            'data_patterns': [],
            'column_suggestions': [],
            'delimiters': Counter(),
            'line_lengths': [],
            'content_samples': []
        }
        
        for file_info in files[:5]:  # Analyze first 5 files
            file_path = Path(file_info['path'])
            analysis = self._analyze_single_ascii_file(file_path)
            ascii_analysis['file_details'].append(analysis)
            
            # Aggregate patterns
            for pattern in analysis.get('patterns', []):
                ascii_analysis['common_patterns'][pattern] += 1
            
            # Sample content
            if analysis.get('sample_content'):
                ascii_analysis['content_samples'].append({
                    'file': file_info['name'],
                    'sample': analysis['sample_content'][:500]
                })
        
        self.analysis_results['ascii'] = ascii_analysis
    
    def _analyze_single_ascii_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze a single ASCII file"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
            
            analysis = {
                'file_name': file_path.name,
                'total_lines': len(lines),
                'patterns': [],
                'delimiters': Counter(),
                'header_lines': [],
                'data_lines': [],
                'sample_content': '',
                'column_structure': {}
            }
            
            # Analyze first 100 lines for patterns
            sample_lines = lines[:100]
            analysis['sample_content'] = '\n'.join(sample_lines[:10])
            
            # Detect patterns
            for line in sample_lines:
                line = line.strip()
                if not line:
                    continue
                
                # Detect delimiters
                delimiters = re.findall(r'[,\t\s]+', line)
                for delim in delimiters:
                    analysis['delimiters'][delim] += 1
                
                # Detect patterns
                if re.match(r'^[A-Z\s]+:', line):
                    analysis['patterns'].append('header_key_value')
                    analysis['header_lines'].append(line)
                elif re.match(r'^[A-Z\s]+$', line):
                    analysis['patterns'].append('header_section')
                elif re.match(r'^[\d\s\.\-]+$', line):
                    analysis['patterns'].append('numeric_data')
                    analysis['data_lines'].append(line)
                elif re.match(r'^[A-Za-z\s]+$', line):
                    analysis['patterns'].append('text_data')
                elif re.match(r'^[A-Za-z0-9\s\.\-]+$', line):
                    analysis['patterns'].append('mixed_data')
                    analysis['data_lines'].append(line)
            
            # Analyze column structure if data lines found
            if analysis['data_lines']:
                analysis['column_structure'] = self._analyze_column_structure(analysis['data_lines'])
            
            return analysis
            
        except Exception as e:
            return {
                'file_name': file_path.name,
                'error': str(e),
                'patterns': ['error']
            }
    
    def _analyze_column_structure(self, data_lines: List[str]) -> Dict[str, Any]:
        """Analyze column structure from data lines"""
        if not data_lines:
            return {}
        
        # Split first few data lines to understand structure
        sample_data = []
        for line in data_lines[:5]:
            # Try different delimiters
            for delimiter in ['\t', ',', '  ', ' ']:
                parts = line.split(delimiter)
                if len(parts) > 1:
                    sample_data.append(parts)
                    break
        
        if not sample_data:
            return {}
        
        # Analyze column structure
        num_columns = len(sample_data[0])
        column_analysis = {
            'num_columns': num_columns,
            'column_types': [],
            'column_samples': [],
            'suggested_names': []
        }
        
        # Analyze each column
        for col_idx in range(num_columns):
            col_values = [row[col_idx] if col_idx < len(row) else '' for row in sample_data]
            
            # Determine column type
            numeric_count = sum(1 for val in col_values if re.match(r'^[\d\.\-]+$', val.strip()))
            if numeric_count == len(col_values):
                col_type = 'numeric'
            elif numeric_count > len(col_values) / 2:
                col_type = 'mixed'
            else:
                col_type = 'text'
            
            column_analysis['column_types'].append(col_type)
            column_analysis['column_samples'].append(col_values[:3])
            
            # Suggest column name based on content
            if col_type == 'numeric':
                if any('depth' in val.lower() for val in col_values):
                    column_analysis['suggested_names'].append('depth')
                elif any('time' in val.lower() for val in col_values):
                    column_analysis['suggested_names'].append('time')
                else:
                    column_analysis['suggested_names'].append(f'col_{col_idx+1}')
            else:
                column_analysis['suggested_names'].append(f'col_{col_idx+1}')
        
        return column_analysis
    
    def _analyze_las_files(self, files: List[Dict]):
        """Analyze LAS files structure"""
        print("  📊 Analyzing LAS file structures...")
        
        las_analysis = {
            'file_type': 'las',
            'total_files': len(files),
            'file_details': [],
            'common_curves': Counter(),
            'well_info_patterns': [],
            'curve_info_patterns': []
        }
        
        for file_info in files[:3]:  # Analyze first 3 files
            file_path = Path(file_info['path'])
            analysis = self._analyze_single_las_file(file_path)
            las_analysis['file_details'].append(analysis)
            
            # Aggregate curve information
            for curve in analysis.get('curves', []):
                las_analysis['common_curves'][curve] += 1
        
        self.analysis_results['las'] = las_analysis
    
    def _analyze_single_las_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze a single LAS file"""
        try:
            analysis = {
                'file_name': file_path.name,
                'curves': [],
                'well_info': {},
                'curve_info': {},
                'data_samples': {},
                'depth_range': {},
                'error': None
            }
            
            # Try to read with lasio
            try:
                las = lasio.read(file_path)
                
                # Extract well information
                if hasattr(las, 'well'):
                    for key, value in las.well.items():
                        if hasattr(value, 'value'):
                            analysis['well_info'][key] = str(value.value)
                
                # Extract curve information
                for curve in las.curves:
                    analysis['curves'].append(curve.mnemonic)
                    analysis['curve_info'][curve.mnemonic] = {
                        'unit': curve.unit,
                        'description': curve.descr,
                        'api_code': curve.api_code
                    }
                
                # Extract data samples
                if hasattr(las, 'df') and not las.df().empty:
                    df = las.df()
                    analysis['depth_range'] = {
                        'start': df.index.min() if not df.empty else None,
                        'stop': df.index.max() if not df.empty else None,
                        'step': df.index[1] - df.index[0] if len(df.index) > 1 else None
                    }
                    
                    # Sample data from each curve
                    for col in df.columns[:5]:  # First 5 curves
                        analysis['data_samples'][col] = df[col].dropna().head(3).tolist()
                
            except Exception as e:
                analysis['error'] = f"lasio error: {str(e)}"
                
                # Fallback: analyze as text file
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                
                # Extract basic LAS structure
                sections = re.findall(r'~([A-Z]+)', content)
                analysis['sections_found'] = sections
                
                # Extract curve information from text
                curve_matches = re.findall(r'([A-Z]+)\s*\.([A-Z]*)\s*([^:]+):', content)
                for match in curve_matches:
                    mnemonic, unit, description = match
                    analysis['curves'].append(mnemonic.strip())
            
            return analysis
            
        except Exception as e:
            return {
                'file_name': file_path.name,
                'error': str(e),
                'curves': []
            }
    
    def _analyze_dlis_files(self, files: List[Dict]):
        """Analyze DLIS files structure"""
        print("  🔧 Analyzing DLIS file structures...")
        
        dlis_analysis = {
            'file_type': 'dlis',
            'total_files': len(files),
            'file_details': [],
            'common_curves': Counter(),
            'file_sizes': []
        }
        
        for file_info in files[:3]:  # Analyze first 3 files
            file_path = Path(file_info['path'])
            analysis = self._analyze_single_dlis_file(file_path)
            dlis_analysis['file_details'].append(analysis)
            dlis_analysis['file_sizes'].append(file_info['size_mb'])
        
        self.analysis_results['dlis'] = dlis_analysis
    
    def _analyze_single_dlis_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze a single DLIS file"""
        try:
            analysis = {
                'file_name': file_path.name,
                'file_size_mb': file_path.stat().st_size / (1024 * 1024),
                'curves': [],
                'frames': [],
                'error': None
            }
            
            # Try to read with dlisio
            try:
                with dlisio.open(file_path) as f:
                    # Extract basic information
                    analysis['frames'] = [frame.name for frame in f.frames]
                    
                    # Extract curve information from frames
                    for frame in f.frames:
                        for curve in frame.curves:
                            analysis['curves'].append(curve.name)
            
            except Exception as e:
                analysis['error'] = f"dlisio error: {str(e)}"
                
                # Fallback: analyze file metadata
                analysis['file_metadata'] = {
                    'size_bytes': file_path.stat().st_size,
                    'created': datetime.fromtimestamp(file_path.stat().st_ctime),
                    'modified': datetime.fromtimestamp(file_path.stat().st_mtime)
                }
            
            return analysis
            
        except Exception as e:
            return {
                'file_name': file_path.name,
                'error': str(e),
                'curves': []
            }
    
    def _analyze_segy_files(self, files: List[Dict]):
        """Analyze SEG-Y files structure"""
        print("  🌊 Analyzing SEG-Y file structures...")
        
        segy_analysis = {
            'file_type': 'segy',
            'total_files': len(files),
            'file_details': [],
            'total_size_gb': sum(f['size_mb'] for f in files) / 1024
        }
        
        for file_info in files[:2]:  # Analyze first 2 files (they're large)
            file_path = Path(file_info['path'])
            analysis = self._analyze_single_segy_file(file_path)
            segy_analysis['file_details'].append(analysis)
        
        self.analysis_results['segy'] = segy_analysis
    
    def _analyze_single_segy_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze a single SEG-Y file"""
        try:
            analysis = {
                'file_name': file_path.name,
                'file_size_gb': file_path.stat().st_size / (1024**3),
                'header_info': {},
                'error': None
            }
            
            # Try to read SEG-Y header
            try:
                with open(file_path, 'rb') as f:
                    # Read text header (first 3200 bytes)
                    text_header = f.read(3200)
                    analysis['header_info']['text_header'] = text_header.decode('ascii', errors='ignore')[:200]
                    
                    # Read binary header (next 400 bytes)
                    binary_header = f.read(400)
                    
                    # Extract basic information from filename
                    filename_parts = file_path.name.split('.')
                    analysis['header_info']['filename_analysis'] = {
                        'parts': filename_parts,
                        'suggested_type': '3D' if '3D' in file_path.name else '2D',
                        'suggested_migration': 'PSDM' if 'PSDM' in file_path.name else 'Unknown',
                        'suggested_algorithm': 'KIRCH' if 'KIRCH' in file_path.name else 'Unknown'
                    }
            
            except Exception as e:
                analysis['error'] = f"SEG-Y reading error: {str(e)}"
            
            return analysis
            
        except Exception as e:
            return {
                'file_name': file_path.name,
                'error': str(e)
            }
    
    def _analyze_pdf_files(self, files: List[Dict]):
        """Analyze PDF files structure"""
        print("  📄 Analyzing PDF file structures...")
        
        pdf_analysis = {
            'file_type': 'pdf',
            'total_files': len(files),
            'file_details': [],
            'content_types': Counter()
        }
        
        for file_info in files[:3]:  # Analyze first 3 files
            file_path = Path(file_info['path'])
            analysis = self._analyze_single_pdf_file(file_path)
            pdf_analysis['file_details'].append(analysis)
            
            if analysis.get('content_type'):
                pdf_analysis['content_types'][analysis['content_type']] += 1
        
        self.analysis_results['pdf'] = pdf_analysis
    
    def _analyze_single_pdf_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze a single PDF file"""
        try:
            analysis = {
                'file_name': file_path.name,
                'content_type': 'unknown',
                'page_count': 0,
                'text_sample': '',
                'error': None
            }
            
            # Try to read PDF
            try:
                with open(file_path, 'rb') as f:
                    pdf_reader = PyPDF2.PdfReader(f)
                    analysis['page_count'] = len(pdf_reader.pages)
                    
                    # Extract text from first page
                    if pdf_reader.pages:
                        first_page = pdf_reader.pages[0]
                        text = first_page.extract_text()
                        analysis['text_sample'] = text[:500]
                        
                        # Determine content type
                        text_lower = text.lower()
                        if any(word in text_lower for word in ['petrophysical', 'interpretation', 'analysis']):
                            analysis['content_type'] = 'petrophysical_report'
                        elif any(word in text_lower for word in ['well', 'borehole', 'trajectory']):
                            analysis['content_type'] = 'well_report'
                        elif any(word in text_lower for word in ['seismic', 'migration', 'reflection']):
                            analysis['content_type'] = 'seismic_report'
                        else:
                            analysis['content_type'] = 'general_document'
            
            except Exception as e:
                analysis['error'] = f"PDF reading error: {str(e)}"
            
            return analysis
            
        except Exception as e:
            return {
                'file_name': file_path.name,
                'error': str(e)
            }
    
    def _analyze_excel_files(self, files: List[Dict]):
        """Analyze Excel files structure"""
        print("  📊 Analyzing Excel file structures...")
        
        excel_analysis = {
            'file_type': 'excel',
            'total_files': len(files),
            'file_details': [],
            'sheet_types': Counter()
        }
        
        for file_info in files[:3]:  # Analyze first 3 files
            file_path = Path(file_info['path'])
            analysis = self._analyze_single_excel_file(file_path)
            excel_analysis['file_details'].append(analysis)
            
            for sheet_type in analysis.get('sheet_types', []):
                excel_analysis['sheet_types'][sheet_type] += 1
        
        self.analysis_results['excel'] = excel_analysis
    
    def _analyze_single_excel_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze a single Excel file"""
        try:
            analysis = {
                'file_name': file_path.name,
                'sheets': [],
                'sheet_types': [],
                'data_samples': {},
                'error': None
            }
            
            # Try to read Excel file
            try:
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_info = {
                        'name': sheet_name,
                        'max_row': sheet.max_row,
                        'max_column': sheet.max_column,
                        'data_type': 'unknown'
                    }
                    
                    # Analyze sheet content
                    if sheet.max_row > 0 and sheet.max_column > 0:
                        # Check first few cells for data type
                        first_cell = sheet.cell(row=1, column=1).value
                        if first_cell and isinstance(first_cell, str):
                            if any(word in first_cell.lower() for word in ['depth', 'md', 'tvd']):
                                sheet_info['data_type'] = 'well_data'
                            elif any(word in first_cell.lower() for word in ['formation', 'layer']):
                                sheet_info['data_type'] = 'geological_data'
                            elif any(word in first_cell.lower() for word in ['porosity', 'permeability']):
                                sheet_info['data_type'] = 'petrophysical_data'
                            else:
                                sheet_info['data_type'] = 'general_data'
                    
                    analysis['sheets'].append(sheet_info)
                    analysis['sheet_types'].append(sheet_info['data_type'])
                
                workbook.close()
            
            except Exception as e:
                analysis['error'] = f"Excel reading error: {str(e)}"
            
            return analysis
            
        except Exception as e:
            return {
                'file_name': file_path.name,
                'error': str(e)
            }
    
    def _analyze_image_files(self, files: List[Dict]):
        """Analyze image files structure"""
        print("  🖼️ Analyzing image file structures...")
        
        image_analysis = {
            'file_type': 'image',
            'total_files': len(files),
            'file_details': [],
            'image_types': Counter()
        }
        
        for file_info in files[:5]:  # Analyze first 5 files
            file_path = Path(file_info['path'])
            analysis = self._analyze_single_image_file(file_path)
            image_analysis['file_details'].append(analysis)
            
            if analysis.get('image_type'):
                image_analysis['image_types'][analysis['image_type']] += 1
        
        self.analysis_results['image'] = image_analysis
    
    def _analyze_single_image_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze a single image file"""
        try:
            analysis = {
                'file_name': file_path.name,
                'image_type': 'unknown',
                'dimensions': None,
                'file_size_mb': file_path.stat().st_size / (1024 * 1024),
                'error': None
            }
            
            # Try to read image
            try:
                with Image.open(file_path) as img:
                    analysis['dimensions'] = img.size
                    analysis['format'] = img.format
                    analysis['mode'] = img.mode
                    
                    # Determine image type based on filename and content
                    filename_lower = file_path.name.lower()
                    if any(word in filename_lower for word in ['log', 'curve', 'well']):
                        analysis['image_type'] = 'well_log'
                    elif any(word in filename_lower for word in ['seismic', 'section', 'migration']):
                        analysis['image_type'] = 'seismic_section'
                    elif any(word in filename_lower for word in ['chart', 'plot', 'graph']):
                        analysis['image_type'] = 'chart_plot'
                    else:
                        analysis['image_type'] = 'general_image'
            
            except Exception as e:
                analysis['error'] = f"Image reading error: {str(e)}"
            
            return analysis
            
        except Exception as e:
            return {
                'file_name': file_path.name,
                'error': str(e)
            }
    
    def _analyze_text_files(self, files: List[Dict]):
        """Analyze text files structure"""
        print("  📝 Analyzing text file structures...")
        
        text_analysis = {
            'file_type': 'text',
            'total_files': len(files),
            'file_details': [],
            'content_types': Counter()
        }
        
        for file_info in files[:5]:  # Analyze first 5 files
            file_path = Path(file_info['path'])
            analysis = self._analyze_single_text_file(file_path)
            text_analysis['file_details'].append(analysis)
            
            if analysis.get('content_type'):
                text_analysis['content_types'][analysis['content_type']] += 1
        
        self.analysis_results['text'] = text_analysis
    
    def _analyze_single_text_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze a single text file"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            analysis = {
                'file_name': file_path.name,
                'content_length': len(content),
                'content_type': 'unknown',
                'sample_content': content[:500],
                'line_count': content.count('\n'),
                'error': None
            }
            
            # Determine content type
            content_lower = content.lower()
            if any(word in content_lower for word in ['well', 'borehole', 'trajectory']):
                analysis['content_type'] = 'well_data'
            elif any(word in content_lower for word in ['formation', 'geological', 'strata']):
                analysis['content_type'] = 'geological_data'
            elif any(word in content_lower for word in ['seismic', 'migration', 'reflection']):
                analysis['content_type'] = 'seismic_data'
            elif any(word in content_lower for word in ['report', 'summary', 'analysis']):
                analysis['content_type'] = 'report'
            else:
                analysis['content_type'] = 'general_text'
            
            return analysis
            
        except Exception as e:
            return {
                'file_name': file_path.name,
                'error': str(e)
            }
    
    def _analyze_dat_files(self, files: List[Dict]):
        """Analyze DAT files structure"""
        print("  📊 Analyzing DAT file structures...")
        
        dat_analysis = {
            'file_type': 'dat',
            'total_files': len(files),
            'file_details': [],
            'content_types': Counter()
        }
        
        for file_info in files[:3]:  # Analyze first 3 files
            file_path = Path(file_info['path'])
            analysis = self._analyze_single_dat_file(file_path)
            dat_analysis['file_details'].append(analysis)
            
            if analysis.get('content_type'):
                dat_analysis['content_types'][analysis['content_type']] += 1
        
        self.analysis_results['dat'] = dat_analysis
    
    def _analyze_single_dat_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze a single DAT file"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            analysis = {
                'file_name': file_path.name,
                'content_length': len(content),
                'content_type': 'unknown',
                'sample_content': content[:500],
                'line_count': content.count('\n'),
                'structure_analysis': {},
                'error': None
            }
            
            # Analyze structure
            lines = content.split('\n')
            analysis['structure_analysis'] = self._analyze_ascii_structure(lines)
            
            # Determine content type
            content_lower = content.lower()
            if 'well' in content_lower and 'picks' in content_lower:
                analysis['content_type'] = 'well_picks'
            elif any(word in content_lower for word in ['formation', 'geological']):
                analysis['content_type'] = 'geological_data'
            else:
                analysis['content_type'] = 'general_data'
            
            return analysis
            
        except Exception as e:
            return {
                'file_name': file_path.name,
                'error': str(e)
            }
    
    def _analyze_ascii_structure(self, lines: List[str]) -> Dict[str, Any]:
        """Analyze ASCII structure from lines"""
        structure = {
            'header_sections': [],
            'data_sections': [],
            'delimiters': Counter(),
            'patterns': Counter()
        }
        
        for line in lines[:100]:  # Analyze first 100 lines
            line = line.strip()
            if not line:
                continue
            
            # Detect delimiters
            delimiters = re.findall(r'[,\t\s]+', line)
            for delim in delimiters:
                structure['delimiters'][delim] += 1
            
            # Detect patterns
            if re.match(r'^[A-Z\s]+:', line):
                structure['patterns']['header_key_value'] += 1
                structure['header_sections'].append(line)
            elif re.match(r'^[A-Z\s]+$', line):
                structure['patterns']['header_section'] += 1
            elif re.match(r'^[\d\s\.\-]+$', line):
                structure['patterns']['numeric_data'] += 1
                structure['data_sections'].append(line)
            elif re.match(r'^[A-Za-z\s]+$', line):
                structure['patterns']['text_data'] += 1
            elif re.match(r'^[A-Za-z0-9\s\.\-]+$', line):
                structure['patterns']['mixed_data'] += 1
                structure['data_sections'].append(line)
        
        return structure
    
    def _analyze_no_extension_files(self, files: List[Dict]):
        """Analyze files without extensions"""
        print("  ❓ Analyzing files without extensions...")
        
        no_ext_analysis = {
            'file_type': 'no_extension',
            'total_files': len(files),
            'file_details': [],
            'content_types': Counter()
        }
        
        for file_info in files[:5]:  # Analyze first 5 files
            file_path = Path(file_info['path'])
            analysis = self._analyze_single_no_extension_file(file_path)
            no_ext_analysis['file_details'].append(analysis)
            
            if analysis.get('content_type'):
                no_ext_analysis['content_types'][analysis['content_type']] += 1
        
        self.analysis_results['no_extension'] = no_ext_analysis
    
    def _analyze_single_no_extension_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze a single file without extension"""
        try:
            # Try to read as text first
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                
                analysis = {
                    'file_name': file_path.name,
                    'content_type': 'text_data',
                    'content_length': len(content),
                    'sample_content': content[:500],
                    'structure_analysis': self._analyze_ascii_structure(content.split('\n')),
                    'error': None
                }
                
                # Determine specific content type
                content_lower = content.lower()
                if any(word in content_lower for word in ['well', 'survey', 'trajectory']):
                    analysis['content_type'] = 'well_survey'
                elif any(word in content_lower for word in ['formation', 'geological']):
                    analysis['content_type'] = 'geological_data'
                else:
                    analysis['content_type'] = 'general_text'
                
                return analysis
                
            except UnicodeDecodeError:
                # Try as binary
                analysis = {
                    'file_name': file_path.name,
                    'content_type': 'binary_data',
                    'file_size_mb': file_path.stat().st_size / (1024 * 1024),
                    'error': None
                }
                
                # Try to determine binary type
                try:
                    mime_type = magic.from_file(str(file_path), mime=True)
                    analysis['mime_type'] = mime_type
                except:
                    analysis['mime_type'] = 'unknown'
                
                return analysis
                
        except Exception as e:
            return {
                'file_name': file_path.name,
                'error': str(e)
            }
    
    def _analyze_unknown_files(self, files: List[Dict], ext: str):
        """Analyze unknown file types"""
        print(f"  ❓ Analyzing unknown file type: {ext}")
        
        unknown_analysis = {
            'file_type': f'unknown_{ext[1:]}' if ext else 'unknown',
            'total_files': len(files),
            'file_details': [],
            'mime_types': Counter()
        }
        
        for file_info in files[:3]:  # Analyze first 3 files
            file_path = Path(file_info['path'])
            analysis = {
                'file_name': file_path.name,
                'file_size_mb': file_info['size_mb'],
                'mime_type': file_info['mime_type'],
                'error': None
            }
            
            unknown_analysis['file_details'].append(analysis)
            unknown_analysis['mime_types'][file_info['mime_type']] += 1
        
        self.analysis_results[f'unknown_{ext[1:]}'] = unknown_analysis
    
    def _generate_analysis_report(self) -> Dict[str, Any]:
        """Generate comprehensive analysis report"""
        print("\n📋 Generating analysis report...")
        
        report = {
            'analysis_timestamp': datetime.now().isoformat(),
            'data_path': str(self.data_path),
            'file_inventory_summary': {},
            'file_type_analysis': self.analysis_results,
            'recommendations': {},
            'parser_suggestions': {}
        }
        
        # Generate file inventory summary
        for ext, files in self.file_inventory.items():
            report['file_inventory_summary'][ext] = {
                'count': len(files),
                'total_size_mb': sum(f['size_mb'] for f in files),
                'avg_size_mb': sum(f['size_mb'] for f in files) / len(files) if files else 0
            }
        
        # Generate recommendations
        report['recommendations'] = self._generate_recommendations()
        
        # Generate parser suggestions
        report['parser_suggestions'] = self._generate_parser_suggestions()
        
        return report
    
    def _generate_recommendations(self) -> Dict[str, Any]:
        """Generate recommendations based on analysis"""
        recommendations = {
            'priority_parsers': [],
            'data_quality_issues': [],
            'file_organization': [],
            'performance_considerations': []
        }
        
        # Priority parsers based on file counts
        file_counts = {ext: len(files) for ext, files in self.file_inventory.items()}
        sorted_types = sorted(file_counts.items(), key=lambda x: x[1], reverse=True)
        
        recommendations['priority_parsers'] = [
            {'file_type': ext, 'count': count, 'priority': i+1}
            for i, (ext, count) in enumerate(sorted_types[:10])
        ]
        
        # Data quality issues
        for ext, analysis in self.analysis_results.items():
            if 'error' in str(analysis):
                recommendations['data_quality_issues'].append({
                    'file_type': ext,
                    'issue': 'parsing_errors',
                    'details': 'Some files could not be parsed'
                })
        
        # Performance considerations
        large_files = []
        for ext, files in self.file_inventory.items():
            for file_info in files:
                if file_info['size_mb'] > 100:  # Files larger than 100MB
                    large_files.append({
                        'file': file_info['name'],
                        'size_mb': file_info['size_mb'],
                        'type': ext
                    })
        
        if large_files:
            recommendations['performance_considerations'].append({
                'issue': 'large_files',
                'files': large_files,
                'suggestion': 'Consider streaming or chunked processing for large files'
            })
        
        return recommendations
    
    def _generate_parser_suggestions(self) -> Dict[str, Any]:
        """Generate parser implementation suggestions"""
        suggestions = {
            'parser_implementations': {},
            'file_type_mappings': {},
            'column_name_suggestions': {}
        }
        
        # Generate parser suggestions for each file type
        for ext, analysis in self.analysis_results.items():
            if ext == 'ascii':
                suggestions['parser_implementations']['ascii'] = {
                    'class_name': 'AsciiParser',
                    'methods': ['parse_header', 'parse_data', 'extract_columns'],
                    'dependencies': ['pandas', 're'],
                    'complexity': 'medium'
                }
            elif ext == 'las':
                suggestions['parser_implementations']['las'] = {
                    'class_name': 'LasParser',
                    'methods': ['parse_well_info', 'parse_curves', 'extract_data'],
                    'dependencies': ['lasio'],
                    'complexity': 'low'
                }
            elif ext == 'dlis':
                suggestions['parser_implementations']['dlis'] = {
                    'class_name': 'DlisParser',
                    'methods': ['parse_frames', 'extract_curves', 'parse_metadata'],
                    'dependencies': ['dlisio'],
                    'complexity': 'high'
                }
            elif ext == 'segy':
                suggestions['parser_implementations']['segy'] = {
                    'class_name': 'SegyParser',
                    'methods': ['parse_header', 'extract_metadata'],
                    'dependencies': ['segpy'],
                    'complexity': 'medium'
                }
            elif ext == 'pdf':
                suggestions['parser_implementations']['pdf'] = {
                    'class_name': 'PdfParser',
                    'methods': ['extract_text', 'extract_tables', 'classify_content'],
                    'dependencies': ['PyPDF2', 'pdfplumber'],
                    'complexity': 'medium'
                }
            elif ext == 'excel':
                suggestions['parser_implementations']['excel'] = {
                    'class_name': 'ExcelParser',
                    'methods': ['parse_sheets', 'extract_data', 'identify_sheet_types'],
                    'dependencies': ['openpyxl', 'pandas'],
                    'complexity': 'low'
                }
            elif ext == 'image':
                suggestions['parser_implementations']['image'] = {
                    'class_name': 'ImageParser',
                    'methods': ['extract_metadata', 'classify_content', 'ocr_text'],
                    'dependencies': ['PIL', 'opencv-python'],
                    'complexity': 'medium'
                }
        
        # Generate file type mappings
        suggestions['file_type_mappings'] = {
            '.asc': 'ascii_parser',
            '.las': 'las_parser',
            '.dlis': 'dlis_parser',
            '.segy': 'segy_parser',
            '.sgy': 'segy_parser',
            '.pdf': 'pdf_parser',
            '.xlsx': 'excel_parser',
            '.xls': 'excel_parser',
            '.png': 'image_parser',
            '.jpg': 'image_parser',
            '.jpeg': 'image_parser',
            '.tiff': 'image_parser',
            '.txt': 'text_parser',
            '.dat': 'dat_parser'
        }
        
        # Generate column name suggestions from actual data
        if 'ascii' in self.analysis_results:
            ascii_analysis = self.analysis_results['ascii']
            for file_detail in ascii_analysis.get('file_details', []):
                if 'column_structure' in file_detail:
                    suggestions['column_name_suggestions'][file_detail['file_name']] = \
                        file_detail['column_structure'].get('suggested_names', [])
        
        return suggestions
    
    def _save_analysis_results(self, report: Dict[str, Any]):
        """Save analysis results to files"""
        print("💾 Saving analysis results...")
        
        # Save full report
        with open('data_analysis_report.json', 'w') as f:
            json.dump(report, f, indent=2, default=str)
        
        # Save summary
        summary = {
            'timestamp': report['analysis_timestamp'],
            'total_file_types': len(report['file_inventory_summary']),
            'total_files': sum(info['count'] for info in report['file_inventory_summary'].values()),
            'total_size_mb': sum(info['total_size_mb'] for info in report['file_inventory_summary'].values()),
            'file_type_breakdown': report['file_inventory_summary'],
            'priority_parsers': report['recommendations']['priority_parsers'][:5]
        }
        
        with open('data_analysis_summary.json', 'w') as f:
            json.dump(summary, f, indent=2, default=str)
        
        print("✅ Analysis results saved to:")
        print("   - data_analysis_report.json (full report)")
        print("   - data_analysis_summary.json (summary)")

if __name__ == "__main__":
    # Run the analysis
    analyzer = DataStructureAnalyzer("data")
    report = analyzer.analyze_all_files()
    
    print("\n🎉 Data structure analysis completed!")
    print(f"📊 Analyzed {len(report['file_inventory_summary'])} file types")
    print(f"📁 Total files: {sum(info['count'] for info in report['file_inventory_summary'].values())}")
    print(f"💾 Total size: {sum(info['total_size_mb'] for info in report['file_inventory_summary'].values()):.1f} MB")
    
    print("\n📋 Top file types found:")
    for i, parser_info in enumerate(report['recommendations']['priority_parsers'][:5], 1):
        print(f"   {i}. {parser_info['file_type']}: {parser_info['count']} files")
    
    print("\n🚀 Ready for Phase 2: Modular Parser Implementation!") 